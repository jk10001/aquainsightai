from __future__ import annotations

import csv
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, Tuple

import openpyxl


@dataclass
class SanitizeResult:
    file_path: str
    file_type: str  # "csv" | "xlsx" | "xlsm"
    replaced_cells: int
    replaced_occurrences: int
    unit_columns_found: int
    modified: bool
    quick_scan_hit: bool


# --- Rules -------------------------------------------------------------

# Replace in ANY column whose header contains "unit" (case-insensitive).
UNIT_COL_RE = re.compile(r"\bunit(s)?\b", re.IGNORECASE)

# Replace µ/μ only when acting like an SI micro-prefix at the start of a token,
# and avoid converting standalone statistical μ such as "μ = ...".
MICRO_PREFIX_RE = re.compile(r"(?<!\w)(?:Âµ|µ|μ)(?=(?:[A-Za-zΩ]))(?!\s*=)")

# Common mojibake for Greek mu: "Î¼" (UTF-8 decoded as cp1252).
MOJIBAKE_GREEK_MU_RE = re.compile(r"(?<!\w)Î¼(?=(?:[A-Za-zΩ]))(?!\s*=)")

# In "unit" columns, replace any occurrences of these characters.
UNIT_COL_FULL_REPL = re.compile(r"(?:Âµ|µ|μ|Î¼)")


def _sanitize_text_value(value: str, aggressive: bool) -> Tuple[str, int]:
    """
    Returns (new_value, num_replacements).
    If aggressive=True: replace any µ/μ occurrences (and common mojibake) with 'u'
    If aggressive=False: replace only typical micro-prefix uses.
    """
    if not value:
        return value, 0

    original = value

    if aggressive:
        value2 = UNIT_COL_FULL_REPL.sub("u", original)
        if value2 == original:
            return original, 0
        n = len(UNIT_COL_FULL_REPL.findall(original))
        return value2, n

    value2 = MICRO_PREFIX_RE.sub("u", original)
    value2 = MOJIBAKE_GREEK_MU_RE.sub("u", value2)
    if value2 == original:
        return original, 0
    n = len(MICRO_PREFIX_RE.findall(original)) + len(
        MOJIBAKE_GREEK_MU_RE.findall(original)
    )
    return value2, n


# --- Quick scan (fast path) -------------------------------------------

# UTF-8 byte sequences for target characters/strings:
#   µ (U+00B5) => C2 B5
#   μ (U+03BC) => CE BC
#   "Â" (U+00C2) => C3 82
#   "Î¼" (U+00CE U+00BC) => C3 8E C2 BC
_QUICK_NEEDLES = (
    b"\xc2\xb5",  # µ
    b"\xce\xbc",  # μ
    b"\xc3\x82",  # Â
    b"\xc3\x8e\xc2\xbc",  # Î¼
)


def _quick_scan_csv(path: Path) -> bool:
    raw = path.read_bytes()
    return any(n in raw for n in _QUICK_NEEDLES)


def _quick_scan_xlsx_like(path: Path) -> bool:
    # Scan relevant XML parts inside the zip without loading openpyxl.
    with zipfile.ZipFile(path, "r") as z:
        for name in z.namelist():
            if not (name.startswith("xl/") and name.endswith(".xml")):
                continue
            data = z.read(name)
            if any(n in data for n in _QUICK_NEEDLES):
                return True
    return False


def _quick_scan(path: Path) -> bool:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _quick_scan_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _quick_scan_xlsx_like(path)
    raise ValueError(f"Unsupported file type '{ext}'. Expected .csv, .xlsx, or .xlsm.")


# --- CSV handling ------------------------------------------------------


# Can this function be replaced with safe_read_csv() in data_description.py?
def _detect_text_encoding(raw: bytes) -> str:
    # Preserve as much as possible without external deps:
    # prefer BOM-aware UTF-8, else UTF-8, else cp1252.
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        raw.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "cp1252"


def _sanitize_csv_inplace(path: Path, quick_scan_hit: bool) -> SanitizeResult:
    raw = path.read_bytes()
    enc = _detect_text_encoding(raw)
    text = raw.decode(enc)

    sample = text[:100_000]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    rows = list(csv.reader(text.splitlines(), dialect))
    if not rows:
        return SanitizeResult(
            str(path), "csv", 0, 0, 0, modified=False, quick_scan_hit=quick_scan_hit
        )

    header = rows[0]
    unit_col_idxs = [
        i
        for i, h in enumerate(header)
        if isinstance(h, str) and UNIT_COL_RE.search(h or "")
    ]
    unit_cols_found = len(unit_col_idxs)

    replaced_cells = 0
    replaced_occ = 0

    # Sanitize header
    new_header = []
    for i, h in enumerate(header):
        if not isinstance(h, str):
            new_header.append(h)
            continue
        aggressive = i in unit_col_idxs or bool(UNIT_COL_RE.search(h))
        new_h, n = _sanitize_text_value(h, aggressive=aggressive)
        new_header.append(new_h)
        if n:
            replaced_cells += 1
            replaced_occ += n
    rows[0] = new_header

    # Sanitize data rows
    for r in range(1, len(rows)):
        row = rows[r]
        for c in range(len(row)):
            v = row[c]
            if not isinstance(v, str):
                continue
            aggressive = c in unit_col_idxs
            new_v, n = _sanitize_text_value(v, aggressive=aggressive)
            if n:
                row[c] = new_v
                replaced_cells += 1
                replaced_occ += n

    modified = replaced_occ > 0

    # SAVE ONLY IF MODIFIED (guard)
    if modified:
        with open(path, "w", encoding=enc, newline="\n") as f:
            writer = csv.writer(
                f,
                dialect=dialect,
                lineterminator="\n",
            )
            writer.writerows(rows)

    return SanitizeResult(
        str(path),
        "csv",
        replaced_cells,
        replaced_occ,
        unit_cols_found,
        modified=modified,
        quick_scan_hit=quick_scan_hit,
    )


# --- XLSX handling -----------------------------------------------------


def _first_nonempty_row(
    values_2d: Iterable[Iterable[object]],
) -> Optional[Tuple[int, list]]:
    """
    Find a header-like row: first row with at least one non-empty cell.
    Returns (1-based row_index, row_values)
    """
    for idx, row in enumerate(values_2d, start=1):
        row_list = list(row)
        if any((v is not None and str(v).strip() != "") for v in row_list):
            return idx, row_list
    return None


def _sanitize_xlsx_like_inplace(path: Path, quick_scan_hit: bool) -> SanitizeResult:
    ext = path.suffix.lower().lstrip(".")
    keep_vba = ext == "xlsm"
    wb = openpyxl.load_workbook(path, keep_vba=keep_vba, data_only=False)

    replaced_cells = 0
    replaced_occ = 0
    unit_cols_found_total = 0
    modified = False

    for ws in wb.worksheets:
        header_info = _first_nonempty_row(ws.iter_rows(values_only=True))
        if header_info is None:
            continue
        header_row_idx, header_row_values = header_info

        unit_col_idxs_1based = []
        for col_idx_1based, h in enumerate(header_row_values, start=1):
            h_str = "" if h is None else str(h)
            if UNIT_COL_RE.search(h_str):
                unit_col_idxs_1based.append(col_idx_1based)

        unit_cols_found_total += len(unit_col_idxs_1based)

        # Sanitize header cells
        for col_idx_1based in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx_1based)
            if isinstance(cell.value, str):
                # Avoid formulas
                if cell.data_type == "f" or cell.value.startswith("="):
                    continue
                aggressive = col_idx_1based in unit_col_idxs_1based or bool(
                    UNIT_COL_RE.search(cell.value)
                )
                new_v, n = _sanitize_text_value(cell.value, aggressive=aggressive)
                if n:
                    cell.value = new_v
                    replaced_cells += 1
                    replaced_occ += n
                    modified = True

        # Sanitize data cells
        for row in ws.iter_rows(
            min_row=header_row_idx + 1, max_row=ws.max_row, max_col=ws.max_column
        ):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                # Avoid formulas
                if cell.data_type == "f" or cell.value.startswith("="):
                    continue
                aggressive = cell.column in unit_col_idxs_1based
                new_v, n = _sanitize_text_value(cell.value, aggressive=aggressive)
                if n:
                    cell.value = new_v
                    replaced_cells += 1
                    replaced_occ += n
                    modified = True

    # SAVE ONLY IF MODIFIED (guard)
    if modified:
        wb.save(path)

    return SanitizeResult(
        str(path),
        ext,
        replaced_cells,
        replaced_occ,
        unit_cols_found_total,
        modified=modified,
        quick_scan_hit=quick_scan_hit,
    )


# --- Public function ---------------------------------------------------


def sanitize_mu_micro_inplace(file_name: str) -> SanitizeResult:
    """
    Programmatic API:

    1) Quick scan the file (fast) for micro/mu characters (and common mojibake).
    2) Only if quick scan hits, run the detailed pass that:
        - converts µ/μ to ASCII 'u' in "unit" columns (aggressive)
        - converts µ/μ to ASCII 'u' in typical unit-prefix contexts elsewhere (conservative)
       and overwrites the file ONLY if changes were made.

    Returns SanitizeResult (including quick_scan_hit + modified).
    """
    path = Path(file_name)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_name}")

    ext = path.suffix.lower()
    if ext not in {".csv", ".xlsx", ".xlsm"}:
        raise ValueError(
            f"Unsupported file type '{ext}'. Expected .csv, .xlsx, or .xlsm."
        )

    quick_hit = _quick_scan(path)

    # If quick scan finds nothing, skip full parsing and do not touch the file.
    if not quick_hit:
        return SanitizeResult(
            file_path=str(path),
            file_type=ext.lstrip("."),
            replaced_cells=0,
            replaced_occurrences=0,
            unit_columns_found=0,
            modified=False,
            quick_scan_hit=False,
        )

    # Otherwise do the slow, structured pass.
    if ext == ".csv":
        return _sanitize_csv_inplace(path, quick_scan_hit=True)
    return _sanitize_xlsx_like_inplace(path, quick_scan_hit=True)


if __name__ == "__main__":
    # Optional CLI usage (kept for convenience; safe to ignore programmatically).
    import sys

    if len(sys.argv) != 2:
        print(
            "Usage: python sanitize_mu_units_inplace.py <file.csv|file.xlsx|file.xlsm>"
        )
        raise SystemExit(2)

    result = sanitize_mu_micro_inplace(sys.argv[1])
    print(result)
